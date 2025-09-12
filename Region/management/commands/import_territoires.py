# core/management/commands/import_territoires.py
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.apps import apps
from pathlib import Path
from openpyxl import load_workbook

from Region.models import Region
from Departement.models import Departement
from Commune.models import Commune


def _clean(val):
    if val is None:
        return None
    s = str(val).strip()
    while "  " in s:
        s = s.replace("  ", " ")
    return s


class Command(BaseCommand):
    help = (
        "Impute Region, Departement, Commune depuis un Excel (colonnes: list_name, name, region, departement, Pays). "
        "N’importe que les lignes correspondant au pays demandé (--pays)."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--pays",
            required=True,
            help="Nom du pays à importer (valeur telle que dans la colonne 'Pays' de l'Excel).",
        )
        parser.add_argument(
            "--file",
            dest="xlsx_path",
            default=None,
            help="Chemin vers l’Excel (défaut: 'modele_data_countries.xlsx' au même niveau que Region/models.py).",
        )
        parser.add_argument(
            "--sheet",
            dest="sheet_name",
            default="Feuil1",
            help="Nom de la feuille Excel (défaut: Feuil1).",
        )
        parser.add_argument(
            "--user-id",
            dest="user_id",
            type=int,
            default=None,
            help="Optionnel: définit le user propriétaire des objets créés.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Ne rien écrire en base, afficher seulement les compteurs.",
        )

    def handle(self, *args, **options):
        # Résoudre chemin du fichier
        if options["xlsx_path"]:
            xlsx_path = Path(options["xlsx_path"])
        else:
            region_app_path = Path(apps.get_app_config("Region").path)
            xlsx_path = region_app_path / "modele_data_countries.xlsx"

        if not xlsx_path.exists():
            raise CommandError(f"Fichier introuvable: {xlsx_path}")

        sheet_name = options["sheet_name"]
        target_country_raw = options["pays"]
        target_country_norm = _clean(target_country_raw).lower() if target_country_raw else None

        user = None
        if options["user_id"]:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            try:
                user = User.objects.get(pk=options["user_id"])
            except User.DoesNotExist:
                raise CommandError(f"User id {options['user_id']} introuvable")

        self.stdout.write(self.style.NOTICE(f"Lecture: {xlsx_path} [{sheet_name}] — Pays ciblé: {target_country_raw}"))

        # Ouvrir Excel
        try:
            wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Impossible d’ouvrir l’Excel: {e}")

        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Feuille '{sheet_name}' absente. Feuilles: {wb.sheetnames}")

        ws = wb[sheet_name]

        # Entêtes
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [(_clean(h) or "") for h in header_row]
        expected = ["list_name", "name", "region", "departement", "Pays"]
        if [h.lower() for h in headers] != [e.lower() for e in expected]:
            raise CommandError(
                f"Entêtes inattendues.\nObtenues: {headers}\nAttendu : {expected}"
            )
        idx = {h: i for i, h in enumerate(headers)}

        created_counts = {"Region": 0, "Departement": 0, "Commune": 0}
        matched_counts = {"Region": 0, "Departement": 0, "Commune": 0}
        errors = 0
        seen_any_for_country = False

        # Contexte d’écriture
        ctx = transaction.atomic if not options["dry_run"] else (lambda: _DummyCtx())

        with ctx():
            for row in ws.iter_rows(min_row=2, values_only=True):
                pays = _clean(row[idx["Pays"]])
                if not pays:
                    continue
                if _clean(pays).lower() != target_country_norm:
                    continue  # pas le pays cible
                seen_any_for_country = True

                list_name = _clean(row[idx["list_name"]])
                name = _clean(row[idx["name"]])
                region_name = _clean(row[idx["region"]])
                dep_name = _clean(row[idx["departement"]])

                if not list_name and not name and not region_name and not dep_name:
                    continue

                try:
                    if list_name == "region":
                        if not name:
                            continue
                        if not options["dry_run"]:
                            obj, created = Region.objects.get_or_create(
                                Nom=name,
                                defaults={"user": user} if user else {},
                            )
                        else:
                            created = True  # simulate
                        created_counts["Region"] += int(created)
                        matched_counts["Region"] += int(not created)

                    elif list_name == "departement":
                        if not name:
                            continue
                        if not region_name:
                            raise ValueError(f"Département '{name}' sans région associée.")
                        if not options["dry_run"]:
                            region, _ = Region.objects.get_or_create(
                                Nom=region_name,
                                defaults={"user": user} if user else {},
                            )
                            dep, created = Departement.objects.get_or_create(
                                Nom=name,
                                Region=region,
                                defaults={"user": user} if user else {},
                            )
                        else:
                            created = True
                        created_counts["Departement"] += int(created)
                        matched_counts["Departement"] += int(not created)

                    elif list_name == "commune":
                        if not name:
                            continue
                        if not dep_name:
                            raise ValueError(f"Commune '{name}' sans département associé.")
                        if not options["dry_run"]:
                            dep_qs = Departement.objects.filter(Nom=dep_name)
                            if not dep_qs.exists():
                                if region_name:
                                    region, _ = Region.objects.get_or_create(
                                        Nom=region_name,
                                        defaults={"user": user} if user else {},
                                    )
                                    dep_obj, _ = Departement.objects.get_or_create(
                                        Nom=dep_name,
                                        Region=region,
                                        defaults={"user": user} if user else {},
                                    )
                                else:
                                    raise ValueError(
                                        f"Département '{dep_name}' introuvable pour la commune '{name}'. "
                                        f"Ajouter la région dans l’Excel ou importer d’abord les départements."
                                    )
                            else:
                                dep_obj = dep_qs.first()

                            com, created = Commune.objects.get_or_create(
                                Nom=name,
                                DepartementID=dep_obj,
                                defaults={"user": user} if user else {},
                            )
                        else:
                            created = True
                        created_counts["Commune"] += int(created)
                        matched_counts["Commune"] += int(not created)

                    else:
                        # on ignore les autres list_name
                        continue

                except Exception as e:
                    errors += 1
                    self.stderr.write(self.style.WARNING(
                        f"[IGNORÉ] pays={pays} {list_name=} {name=} {region_name=} {dep_name=} -> {e}"
                    ))

        if not seen_any_for_country:
            raise CommandError(
                f"Aucune ligne trouvée pour le pays '{target_country_raw}'. "
                f"Vérifie l’orthographe (exacte à la casse/accents non requis)."
            )

        mode = "DRY-RUN (aucune écriture)" if options["dry_run"] else "ÉCRITURE OK"
        self.stdout.write(self.style.SUCCESS(f"Import ({mode}) terminé pour: {target_country_raw}"))
        self.stdout.write(
            f"Regions    -> créés: {created_counts['Region']}, existants: {matched_counts['Region']}"
        )
        self.stdout.write(
            f"Départements -> créés: {created_counts['Departement']}, existants: {matched_counts['Departement']}"
        )
        self.stdout.write(
            f"Communes   -> créées: {created_counts['Commune']}, existantes: {matched_counts['Commune']}"
        )
        if errors:
            self.stdout.write(self.style.WARNING(f"💡 Lignes ignorées pour erreurs: {errors}"))


class _DummyCtx:
    """Contexte factice pour --dry-run (ne rien écrire)."""
    def __enter__(self): return None
    def __exit__(self, exc_type, exc, tb): return False
