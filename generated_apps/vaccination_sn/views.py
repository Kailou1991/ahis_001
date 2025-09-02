# generated_apps/vaccination_sn/views.py
from __future__ import annotations

# ───────────────────────── stdlib ─────────────────────────
from dataclasses import dataclass
from io import BytesIO
from typing import Iterable, Optional, Sequence
import base64

# ───────────────────────── django ─────────────────────────
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum, F, Q, QuerySet
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.shortcuts import render
from django.views.generic import ListView, TemplateView, View
from django.db.models import Q
import re


# ─────────────────────── third-party ──────────────────────
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# ──────────────────────── local apps ──────────────────────
from .models import (
    VaccinationSn,
    VaccinationSnChild0c8ff1d1 as VChild,
)
from generated_apps.objectif_sn.models import (
    ObjectifSn,
    ObjectifSnChild0c8ff1d1 as OChild,
)
# si nécessaire, adaptez le chemin de l’app “materiel”
from materiel.models import Dotation, DotationDoseVaccin


# ╔════════════════════════════════════════════════════════╗
# ║                     Helpers communs                    ║
# ╚════════════════════════════════════════════════════════╝


from typing import Sequence, Iterable, Optional
from io import BytesIO
import base64, re

from django.db.models.query import QuerySet

# --- IDs de session -> libellés Kobo ---
def _session_region_dept_names(request):
    region_name = depart_name = None
    try:
        rid = request.session.get('region_id')
        did = request.session.get('departement_id')
        if rid:
            from Region.models import Region
            region_name = Region.objects.filter(id=rid).values_list("Nom", flat=True).first()
        if did:
            from Departement.models import Departement
            depart_name = Departement.objects.filter(id=did).values_list("Nom", flat=True).first()
    except Exception:
        pass
    return region_name, depart_name

# --- Restreint n'importe quel QS au périmètre session ---
def _restrict_to_session(qs: QuerySet, request, region_lookup: str, dept_lookup: str) -> QuerySet:
    region_name, depart_name = _session_region_dept_names(request)
    if region_name:
        qs = qs.filter(**{region_lookup: region_name})
    if depart_name:
        qs = qs.filter(**{dept_lookup: depart_name})
    return qs


@dataclass(frozen=True)
class FilterParams:
    campagne: str = ""
    type_campagne: str = ""
    maladie: str = ""
    region: str = ""
    departement: str = ""


MONTH_LABELS: Sequence[str] = (
    "Jan", "Fev", "Mar", "Avr", "Mai", "Juin",
    "Juil", "Aout", "Sept", "Oct", "Nov", "Déc"
)


def parse_filters(request) -> FilterParams:
    """Extrait et nettoie les filtres GET."""
    g = request.GET
    return FilterParams(
        campagne=(g.get("campagne") or "").strip(),
        type_campagne=(g.get("type_campagne") or "").strip(),
        maladie=(g.get("maladie") or "").strip(),
        region=(g.get("region") or "").strip(),
        departement=(g.get("departement") or "").strip(),
    )


def apply_filters_to_vchild(qs: QuerySet, f: FilterParams) -> QuerySet:
    """Applique les filtres aux items Enfants Vaccination."""
    if f.campagne:
        qs = qs.filter(parent__campagne__iexact=f.campagne)
    if f.type_campagne:
        qs = qs.filter(parent__type_de_campagne__iexact=f.type_campagne)
    if f.maladie:
        qs = qs.filter(maladie_masse__icontains=f.maladie)
    if f.region:
        qs = qs.filter(parent__grp4_region__iexact=f.region)
    if f.departement:
        qs = qs.filter(parent__grp4_departement__iexact=f.departement)
    return qs


def apply_filters_to_ochild(qs: QuerySet, f: FilterParams) -> QuerySet:
    """Applique les filtres aux items Enfants Objectif."""
    if f.campagne:
        qs = qs.filter(parent__campagne__iexact=f.campagne)
    if f.type_campagne:
        qs = qs.filter(parent__type_de_campagne__iexact=f.type_campagne)
    if f.maladie:
        qs = qs.filter(maladie_masse__iexact=f.maladie)
    if f.region:
        qs = qs.filter(parent__grp4_region__iexact=f.region)
    if f.departement:
        qs = qs.filter(parent__grp4_departement__iexact=f.departement)
    return qs


def distinct_non_empty(qs: QuerySet, field: str) -> Iterable[str]:
    """Renvoie des valeurs distinctes non vides pour un champ donné."""
    return (
        qs.exclude(**{f"{field}__isnull": True})
          .exclude(**{field: ""})
          .values_list(field, flat=True)
          .distinct()
    )


def generate_graph(
    x: Sequence,
    y: Sequence,
    title: str,
    xlabel: str,
    ylabel: str,
    chart_type: str = "bar",
    y2: Optional[Sequence] = None,
) -> str:
    """Retourne une image Matplotlib encodée en base64."""
    plt.switch_backend("Agg")
    fig, ax = plt.subplots(figsize=(8, 4))

    if chart_type == "bar":
        colors = plt.cm.tab20.colors
        color_list = [colors[i % len(colors)] for i in range(len(x))]
        bars = ax.bar(x, y, color=color_list)
        for bar in bars:
            h = bar.get_height()
            ax.annotate(
                f"{h:,.0f}".replace(",", " "),
                xy=(bar.get_x() + bar.get_width() / 2, h),
                xytext=(0, 3),
                textcoords="offset points",
                ha="center",
                va="bottom",
                fontsize=8,
            )
        ax.yaxis.set_visible(False)

    elif chart_type == "grouped_bar" and y2 is not None:
        bar_width = 0.35
        index = list(range(len(x)))
        colors = plt.cm.tab10.colors
        ax.bar([i - bar_width / 2 for i in index], y, bar_width, label="Objectif", color=colors[0])
        ax.bar([i + bar_width / 2 for i in index], y2, bar_width, label="Vaccinés", color=colors[1])
        ax.set_xticks(index)
        ax.set_xticklabels(x, rotation=45, ha="right")
        for i, val in enumerate(y):
            ax.annotate(
                f"{val:,.0f}".replace(",", " "),
                xy=(i - bar_width / 2, val),
                xytext=(0, 3),
                textcoords="offset points",
                ha="center",
                fontsize=8,
            )
        for i, val in enumerate(y2):
            ax.annotate(
                f"{val:,.0f}".replace(",", " "),
                xy=(i + bar_width / 2, val),
                xytext=(0, 3),
                textcoords="offset points",
                ha="center",
                fontsize=8,
            )
        ax.legend()
        ax.yaxis.set_visible(False)

    elif chart_type == "line":
        ax.plot(x, y, marker="o", linestyle="-")
        for i, val in enumerate(y):
            ax.annotate(
                f"{val:,.0f}".replace(",", " "),
                (x[i], y[i]),
                textcoords="offset points",
                xytext=(0, 5),
                ha="center",
                fontsize=8,
            )

    ax.set_title(title)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    fig.tight_layout()

    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    buf.seek(0)
    png = buf.getvalue()
    buf.close()
    plt.close(fig)

    return base64.b64encode(png).decode("utf-8")


# ╔════════════════════════════════════════════════════════╗
# ║                 Liste simple (statistiques)            ║
# ╚════════════════════════════════════════════════════════╝

class VaccinationStats(ListView):
    """
    Une ligne = un item du repeat (commune/maladie).
    Filtres : campagne, type_de_campagne, maladie, région, département.
    """
    model = VChild
    template_name = "vaccination_sn/stats.html"
    context_object_name = "rows"
    paginate_by = 50

    # Base QS
    def _base_queryset(self) -> QuerySet:
        qs = (
            VChild.objects.select_related("parent")
            .order_by("-parent__submission_time", "-parent_id", "item_index")
        )
        # ⬇️ périmètre session : region/departement des PARENTS de VChild
        qs = _restrict_to_session(qs, self.request,
                                  region_lookup="parent__grp4_region__iexact",
                                  dept_lookup="parent__grp4_departement__iexact")
        return qs
    # Queryset filtré
    def get_queryset(self) -> QuerySet:
        f = parse_filters(self.request)
        return apply_filters_to_vchild(self._base_queryset(), f)

    # Contexte : totaux + listes
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        f = parse_filters(self.request)
        qs = self.get_queryset()

        # Totaux
        agg = qs.aggregate(
            vacc_prive=Sum("vaccine_prive"),
            vacc_public=Sum("vaccine_public"),
            marques=Sum("total_marque"),
        )
        ctx["totals"] = {
            "rows": qs.count(),
            "vacc_prive": agg["vacc_prive"] or 0,
            "vacc_public": agg["vacc_public"] or 0,
            "marques": agg["marques"] or 0,
        }

        base = self._base_queryset()

        # Listes globales
        ctx["campaigns"] = (
            distinct_non_empty(base, "parent__campagne").order_by("parent__campagne")
        )
        ctx["maladies"] = (
            distinct_non_empty(base, "maladie_masse").order_by("maladie_masse")
        )
        ctx["types"] = (
            distinct_non_empty(base, "parent__type_de_campagne").order_by("parent__type_de_campagne")
        )

        # Régions (dépendent de campagne/type/maladie)
        qs_for_regions = apply_filters_to_vchild(base, FilterParams(
            campagne=f.campagne, type_campagne=f.type_campagne, maladie=f.maladie
        ))
        ctx["regions"] = (
            distinct_non_empty(qs_for_regions, "parent__grp4_region").order_by("parent__grp4_region")
        )

        # Départements (dépendent en plus de la région)
        qs_for_deps = apply_filters_to_vchild(base, FilterParams(
            campagne=f.campagne, type_campagne=f.type_campagne, maladie=f.maladie, region=f.region
        ))
        ctx["departements"] = (
            distinct_non_empty(qs_for_deps, "parent__grp4_departement").order_by("parent__grp4_departement")
        )

        ctx["filters"] = {
            "campagne": f.campagne,
            "type_campagne": f.type_campagne,
            "maladie": f.maladie,
            "region": f.region,
            "departement": f.departement,
        }
        return ctx


def _norm_key(s: str) -> str:
        """clé de dédup: trim + collapse espaces + lowercase"""
        s = (s or "").strip()
        s = re.sub(r"\s+", " ", s)
        return s.lower()

def _pretty_label(s: str) -> str:
        """label affiché propre (trim + collapse espaces)"""
        s = (s or "").strip()
        return re.sub(r"\s+", " ", s)


class DepartementsForRegion(View):
        """
        Renvoie les départements disponibles pour la région sélectionnée
        (en tenant compte des autres filtres). Déduplique (casse/espaces) et trie.
        GET params: campagne, type_campagne, maladie, region
        """
        def get(self, request, *args, **kwargs):
            campagne      = (request.GET.get("campagne") or "").strip()
            type_campagne = (request.GET.get("type_campagne") or "").strip()
            maladie       = (request.GET.get("maladie") or "").strip()
            region        = (request.GET.get("region") or "").strip()

            qs = VChild.objects.select_related("parent")
            # ⬇️ périmètre session
            qs = _restrict_to_session(qs, request,
                                    region_lookup="parent__grp4_region__iexact",
                                    dept_lookup="parent__grp4_departement__iexact")
            # puis tes filtres GET existants...
            if campagne:
                qs = qs.filter(parent__campagne__icontains=campagne)
            if type_campagne:
                qs = qs.filter(parent__type_de_campagne__iexact=type_campagne)
            if maladie:
                qs = qs.filter(maladie_masse__icontains=maladie)
            if region:
                qs = qs.filter(parent__grp4_region__iexact=region)

            # Récup brut
            raw = list(
                qs.exclude(parent__grp4_departement__isnull=True)
                .exclude(parent__grp4_departement="")
                .values_list("parent__grp4_departement", flat=True)
            )

            # Déduplication robuste
            seen = {}
            for d in raw:
                key = _norm_key(d)
                if key:  # ignore None/""
                    # garde le premier libellé "propre" rencontré
                    seen.setdefault(key, _pretty_label(d))

            # Tri alpha insensible à la casse
            deps = sorted(seen.values(), key=lambda s: s.casefold())

            return JsonResponse({"departements": deps})

# ╔════════════════════════════════════════════════════════╗
# ║                   Dashboard consolidé                  ║
# ╚════════════════════════════════════════════════════════╝
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin

class KoboDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "vaccination_sn/dashboard.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        f = parse_filters(self.request)

        # Base QS enfants (restreints au périmètre session région/département)
        vbase = _restrict_to_session(
            VChild.objects.select_related("parent"),
            self.request, "parent__grp4_region__iexact", "parent__grp4_departement__iexact"
        )
        obase = _restrict_to_session(
            OChild.objects.select_related("parent"),
            self.request, "parent__grp4_region__iexact", "parent__grp4_departement__iexact"
        )

        # QS après filtres UI
        vqs = apply_filters_to_vchild(vbase, f)
        oqs = apply_filters_to_ochild(obase, f)

        # === Totaux Vaccination ===
        agg_v = vqs.aggregate(
            v_pub=Coalesce(Sum("vaccine_public"), 0),
            v_pri=Coalesce(Sum("vaccine_prive"), 0),
            marques=Coalesce(Sum("total_marque"), 0),
        )
        total_vacc_public = int(agg_v["v_pub"])
        total_vacc_prive  = int(agg_v["v_pri"])
        total_vaccines    = total_vacc_public + total_vacc_prive
        total_marques     = int(agg_v["marques"])

        # === Totaux Objectifs ===
        agg_o = oqs.aggregate(
            obj_cible=Coalesce(Sum("effectif_cible"), 0),
            obj_eligible=Coalesce(Sum("effectif_elligible"), 0),
        )
        total_objectif = int(agg_o["obj_cible"])
        total_eligible = int(agg_o["obj_eligible"])

        taux_realisation = round((total_vaccines / total_objectif) * 100, 2) if total_objectif else 0.0
        taux_couverture  = round((total_vaccines / total_eligible) * 100, 2) if total_eligible else 0.0

        # === Listes (campagnes/types/maladies) – restreindre aussi les PARENTS ===
        vparent = _restrict_to_session(
            VaccinationSn.objects.all(),
            self.request, "grp4_region__iexact", "grp4_departement__iexact"
        )
        oparent = _restrict_to_session(
            ObjectifSn.objects.all(),
            self.request, "grp4_region__iexact", "grp4_departement__iexact"
        )

        camps_v = distinct_non_empty(vparent, "campagne")
        camps_o = distinct_non_empty(oparent, "campagne")
        campaigns = sorted(set(list(camps_v) + list(camps_o)))

        t_v = distinct_non_empty(vparent, "type_de_campagne")
        t_o = distinct_non_empty(oparent, "type_de_campagne")
        types_campagne = sorted(set(list(t_v) + list(t_o)))

        mals_v = distinct_non_empty(vbase, "maladie_masse")
        mals_o = distinct_non_empty(obase, "maladie_masse")
        maladies = sorted(set(list(mals_v) + list(mals_o)))

        # === Régions visibles (dépendent des autres filtres) – rester dans le périmètre ===
        base_for_regions_v = apply_filters_to_vchild(vbase, FilterParams(
            campagne=f.campagne, type_campagne=f.type_campagne, maladie=f.maladie
        ))
        base_for_regions_o = apply_filters_to_ochild(obase, FilterParams(
            campagne=f.campagne, type_campagne=f.type_campagne, maladie=f.maladie
        ))
        regs_v = distinct_non_empty(base_for_regions_v, "parent__grp4_region")
        regs_o = distinct_non_empty(base_for_regions_o, "parent__grp4_region")
        regions = sorted(set(list(regs_v) + list(regs_o)))

        # === Départements (dépendent de la région sélectionnée) ===
        base_for_deps_v = apply_filters_to_vchild(base_for_regions_v, FilterParams(region=f.region))
        base_for_deps_o = apply_filters_to_ochild(base_for_regions_o, FilterParams(region=f.region))
        deps_v = distinct_non_empty(base_for_deps_v, "parent__grp4_departement")
        deps_o = distinct_non_empty(base_for_deps_o, "parent__grp4_departement")
        departements = sorted(set(list(deps_v) + list(deps_o)))

        # === Graphe Objectifs vs Vaccinés par région (barres groupées) ===
        obj_by_reg = (
            oqs.values("parent__grp4_region")
               .annotate(obj=Coalesce(Sum("effectif_cible"), 0))
               .order_by("parent__grp4_region")
        )
        vac_by_reg = (
            vqs.values("parent__grp4_region")
               .annotate(v_pub=Coalesce(Sum("vaccine_public"), 0),
                         v_pri=Coalesce(Sum("vaccine_prive"), 0))
               .order_by("parent__grp4_region")
        )
        vac_reg_map = {r["parent__grp4_region"]: int(r["v_pub"] + r["v_pri"]) for r in vac_by_reg}
        x_regions, y_obj, y_vac = [], [], []
        for r in obj_by_reg:
            reg = r["parent__grp4_region"] or "—"
            x_regions.append(reg)
            y_obj.append(int(r["obj"] or 0))
            y_vac.append(int(vac_reg_map.get(reg, 0)))
        graph_region_grouped = ""
        if x_regions:
            graph_region_grouped = generate_graph(
                x_regions, y_obj, "Objectifs vs Vaccinés par région",
                "Régions", "Effectifs", chart_type="grouped_bar", y2=y_vac
            )

        # === Graphe mensuel (soumissions) – ligne ===
        monthly = (
            vqs.values("parent__submission_time__year", "parent__submission_time__month")
               .annotate(
                   total=Coalesce(Sum("vaccine_public"), 0) + Coalesce(Sum("vaccine_prive"), 0)
               )
               .order_by("parent__submission_time__year", "parent__submission_time__month")
        )
        x_mois, y_mois = [], []
        for m in monthly:
            yy = m["parent__submission_time__year"]
            mm = m["parent__submission_time__month"]
            if yy and mm:
                x_mois.append(f"{MONTH_LABELS[mm-1]}-{yy}")
                y_mois.append(int(m["total"] or 0))
        graph_mensuel = ""
        if x_mois:
            graph_mensuel = generate_graph(
                x_mois, y_mois, "Tendance mensuelle des vaccinations",
                "Mois", "Effectifs", chart_type="line"
            )

        # === Tableau départements ===
        rows_dep = []
        obj_dep = (
            oqs.values("parent__grp4_departement")
               .annotate(obj=Coalesce(Sum("effectif_cible"), 0),
                         elig=Coalesce(Sum("effectif_elligible"), 0))
        )
        vac_dep = (
            vqs.values("parent__grp4_departement")
               .annotate(v_pub=Coalesce(Sum("vaccine_public"), 0),
                         v_pri=Coalesce(Sum("vaccine_prive"), 0),
                         marque=Coalesce(Sum("total_marque"), 0))
        )
        obj_map = {r["parent__grp4_departement"]: (int(r["obj"]), int(r["elig"])) for r in obj_dep}
        vac_map = {r["parent__grp4_departement"]: (int(r["v_pub"] + r["v_pri"]), int(r["marque"])) for r in vac_dep}

        all_deps = set(obj_map) | set(vac_map)
        for dep in sorted(all_deps, key=lambda v: (v is None, (v or "").lower())):
            obj, elig = obj_map.get(dep, (0, 0))
            vac, marq = vac_map.get(dep, (0, 0))
            taux_r = round((vac / obj) * 100, 2) if obj else 0.0
            rows_dep.append({
                "departement": dep or "—",
                "objectif": obj,
                "eligible": elig,
                "vaccines": vac,
                "marques": marq,
                "taux_realisation": taux_r,
            })

        # === Dotations – doses ===
        doses_qs = DotationDoseVaccin.objects.select_related("campagne", "maladie")
        if f.campagne:
            doses_qs = doses_qs.filter(campagne__Campagne__iexact=f.campagne)
        if f.maladie:
            doses_qs = doses_qs.filter(maladie__Maladie__iexact=f.maladie)

        doses_rows = list(
            doses_qs.values("campagne__Campagne", "maladie__Maladie")
                    .annotate(doses=Coalesce(Sum("quantite_doses"), 0))
                    .order_by("campagne__Campagne", "maladie__Maladie")
        )
        total_doses = int(doses_qs.aggregate(t=Coalesce(Sum("quantite_doses"), 0))["t"] or 0)

        doses_by_mal = (
            doses_qs.values("maladie__Maladie")
                    .annotate(total=Coalesce(Sum("quantite_doses"), 0))
                    .order_by("maladie__Maladie")
        )
        x_dm = [r["maladie__Maladie"] for r in doses_by_mal]
        y_dm = [int(r["total"] or 0) for r in doses_by_mal]
        graph_doses_by_maladie = generate_graph(
            x_dm, y_dm, "Dotations (doses) par maladie", "Maladies", "Doses", chart_type="bar"
        ) if x_dm else ""

        # === Dotations – matériel ===
        mat_qs = Dotation.objects.select_related("campagne", "region", "type_materiel")
        if f.campagne:
            mat_qs = mat_qs.filter(campagne__Campagne__iexact=f.campagne)
        if f.region:
            mat_qs = mat_qs.filter(region__Nom__iexact=f.region)

        materiel_rows = list(
            mat_qs.values("campagne__Campagne", "region__Nom", "type_materiel__nom")
                 .annotate(quantite=Coalesce(Sum("quantite"), 0))
                 .order_by("campagne__Campagne", "region__Nom", "type_materiel__nom")
        )
        total_matos = int(mat_qs.aggregate(t=Coalesce(Sum("quantite"), 0))["t"] or 0)

        mat_by_type = (
            mat_qs.values("type_materiel__nom")
                 .annotate(total=Coalesce(Sum("quantite"), 0))
                 .order_by("type_materiel__nom")
        )
        x_mt = [r["type_materiel__nom"] for r in mat_by_type]
        y_mt = [int(r["total"] or 0) for r in mat_by_type]
        graph_matos_by_type = generate_graph(
            x_mt, y_mt, "Dotations matériel par type", "Types", "Quantités", chart_type="bar"
        ) if x_mt else ""

        # === Contexte final ===
        ctx.update({
            "filters": {
                "campagne": f.campagne,
                "type_campagne": f.type_campagne,
                "maladie": f.maladie,
                "region": f.region,
                "departement": f.departement,
            },
            "campaigns": campaigns,
            "types_campagne": types_campagne,
            "maladies": maladies,
            "regions": regions,
            "departements": departements,

            "totals": {
                "objectif": total_objectif,
                "eligible": total_eligible,
                "vacc_public": total_vacc_public,
                "vacc_prive": total_vacc_prive,
                "vacc_total": total_vaccines,
                "marques": total_marques,
                "taux_realisation": taux_realisation,
                "taux_couverture": taux_couverture,
            },

            "graph_region_grouped": graph_region_grouped,
            "graph_mensuel": graph_mensuel,
            "tableau_departements": rows_dep,

            "doses_rows": doses_rows,
            "total_doses": total_doses,
            "graph_doses_by_maladie": graph_doses_by_maladie,

            "materiel_rows": materiel_rows,
            "total_matos": total_matos,
            "graph_matos_by_type": graph_matos_by_type,
        })
        return ctx

    
    ################Rapport campagne################################""

from datetime import date, timedelta
from io import BytesIO
from collections import defaultdict

from django import forms
from django.db.models import Sum, F
from django.db.models.functions import Coalesce, TruncWeek, TruncDate
from django.http import HttpResponse
from django.shortcuts import render


# Excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ----------------------------- Outils génériques -----------------------------
def _distinct_non_empty(qs, field):
    """Liste distincte d'un champ texte en excluant None/''."""
    return (
        qs.exclude(**{f"{field}__isnull": True})
          .exclude(**{field: ""})
          .values_list(field, flat=True)
          .distinct()
    )


def _campaign_choices():
    """
    Construit la liste des campagnes disponibles en union
    (ObjectifSn.campagne ∪ VaccinationSn.campagne), triées.
    """
    camps_o = _distinct_non_empty(ObjectifSn.objects.all(), "campagne")
    camps_v = _distinct_non_empty(VaccinationSn.objects.all(), "campagne")
    return [(c, c) for c in sorted(set(list(camps_o) + list(camps_v)))]


def _shift_year_safe(d: date, years: int) -> date:
    """Décale une date d'un nombre d'années en gérant les 29/02."""
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        # 29/02 -> 28/02 si année cible non bissextile
        return d.replace(month=2, day=28, year=d.year + years)


# ----------------------------- Formulaire de filtre -----------------------------
class RapportCamvacFilterForm(forms.Form):
    PERIODES = (("hebdo", "Hebdomadaire"), ("mensuel", "Mensuel"))

    campagne = forms.ChoiceField(
        choices=[],
        required=True,
        label="Campagne",
        widget=forms.Select(attrs={"class": "form-control"})
    )
    periode = forms.ChoiceField(
        choices=PERIODES,
        initial="hebdo",
        required=True,
        label="Période",
        widget=forms.Select(attrs={"class": "form-control"})
    )

    # Hebdo
    semaine_iso = forms.CharField(
        required=False,
        label="Semaine ISO",
        help_text="ex: 2025-W22",
        widget=forms.TextInput(attrs={"class": "form-control", "placeholder": "2025-W22"})
    )

    # Mensuel (input type="month")
    mois = forms.DateField(
        required=False,
        label="Mois",
        input_formats=["%Y-%m", "%Y-%m-%d"],
        widget=forms.DateInput(attrs={"type": "month", "class": "form-control"})
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["campagne"].choices = _campaign_choices()

    def clean(self):
        cleaned = super().clean()
        periode = cleaned.get("periode")

        if periode == "hebdo":
            if not cleaned.get("semaine_iso"):
                self.add_error("semaine_iso", "La semaine ISO est requise (ex: 2025-W22).")
            cleaned["mois"] = None

        elif periode == "mensuel":
            if not cleaned.get("mois"):
                self.add_error("mois", "Le mois est requis (format AAAA-MM).")
            cleaned["semaine_iso"] = ""

        return cleaned


# ----------------------------- Fenêtre temporelle -----------------------------
def resolve_period(cleaned):
    """
    Retourne (start_date, end_date, label_periode) selon hebdo/mensuel.
    Filtres appliqués sur parent__submission_time (partie date).
    """
    if cleaned["periode"] == "hebdo":
        year, week = cleaned["semaine_iso"].split("-W")
        year, week = int(year), int(week)
        start = date.fromisocalendar(year, week, 1)  # lundi
        end   = date.fromisocalendar(year, week, 7)  # dimanche
        return start, end, f"Semaine {year}-W{week:02d}"
    else:
        d = cleaned["mois"]  # 1er du mois (HTML month picker)
        start = d.replace(day=1)
        if d.month == 12:
            first_next = date(d.year + 1, 1, 1)
        else:
            first_next = date(d.year, d.month + 1, 1)
        end = first_next - date.resolution
        return start, end, d.strftime("%B-%Y")


# ----------------------------- Ecriture Excel -----------------------------
def write_sheets(named_tables: dict[str, list[list]]):
    wb = Workbook()
    first_title = list(named_tables.keys())[0]
    ws0 = wb.active
    ws0.title = first_title

    for title, rows in named_tables.items():
        ws = wb[ws0.title] if ws0.title == title else wb.create_sheet(title)
        for r in rows:
            ws.append(r)
        # colonnes auto
        for col in ws.columns:
            maxlen = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(maxlen + 2, 45))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws0 = ws

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


# ----------------------------- Tables (T1..T9) -----------------------------
def _vsum(qs):
    """Retourne un dict(key -> vaccinés) où vaccinés = somme(public) + somme(prive)."""
    data = (
        qs.annotate(v_pub=Coalesce(Sum("vaccine_public"), 0),
                    v_pri=Coalesce(Sum("vaccine_prive"), 0))
          .values_list("key", "v_pub", "v_pri")
    )
    out = defaultdict(int)
    for k, a, b in data:
        out[k] += int(a or 0) + int(b or 0)
    return out


def t1_situation_regionale(start, end, camp_name: str):
    # Objectifs par (région, maladie)
    obj_qs = (
        OChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name)
        .values("parent__grp4_region", "maladie_masse")
        .annotate(objectif=Coalesce(Sum("effectif_cible"), 0))
    )

    # Vaccinés (submission_time dans [start,end]) par (région, maladie)
    vac_qs = (
        VChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name,
                parent__submission_time__date__range=[start, end])
        .values(key=F("parent__grp4_region") + F("maladie_masse"))  # placeholder pour group key
        # on va regrouper manuellement par tuple (region, maladie) juste après
    )

    # Comme on ne peut pas concaténer deux champs texte directement pour la clé,
    # on recalcule proprement la somme via une double boucle
    vac_rows = (
        VChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name,
                parent__submission_time__date__range=[start, end])
        .values("parent__grp4_region", "maladie_masse")
        .annotate(v_pub=Coalesce(Sum("vaccine_public"), 0),
                  v_pri=Coalesce(Sum("vaccine_prive"), 0))
    )
    vac_map = {(r["parent__grp4_region"], r["maladie_masse"]): int(r["v_pub"] + r["v_pri"]) for r in vac_rows}

    # Doses livrées par maladie (national, faute de détail régional)
    doses = (
        DotationDoseVaccin.objects
        .filter(campagne__Campagne__iexact=camp_name, date_dotation__lte=end)
        .values("maladie__Maladie")
        .annotate(doses=Coalesce(Sum("quantite_doses"), 0))
    )
    doses_dict = {d["maladie__Maladie"]: int(d["doses"] or 0) for d in doses}

    # Maps objectifs
    obj_map = defaultdict(int)
    regions = set()
    maladies = set()
    for o in obj_qs:
        reg = o["parent__grp4_region"] or "—"
        mal = o["maladie_masse"] or "—"
        obj_map[(reg, mal)] += int(o["objectif"] or 0)
        regions.add(reg); maladies.add(mal)

    # Clés (union avec vaccins)
    for (reg, mal), v in vac_map.items():
        regions.add(reg or "—"); maladies.add(mal or "—")

    rows = [["Région", "Maladie", "Objectifs", "Vaccins livrés (doses)", "Effectifs vaccinés", "Écart (Obj - Vacc)"]]
    keys = sorted(obj_map.keys() | vac_map.keys(), key=lambda x: (x[0] or "", x[1] or ""))
    for (reg, mal) in keys:
        o = obj_map.get((reg, mal), 0)
        v = vac_map.get((reg, mal), 0)
        d = doses_dict.get(mal, 0)
        rows.append([reg or "—", mal or "—", o, d, v, (o or 0) - (v or 0)])
    return rows


def t2_ecarts_nationaux(start, end, camp_name: str):
    obj_qs = (
        OChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name)
        .values("maladie_masse")
        .annotate(objectif=Coalesce(Sum("effectif_cible"), 0))
    )
    vac_qs = (
        VChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name,
                parent__submission_time__date__range=[start, end])
        .values("maladie_masse")
        .annotate(v_pub=Coalesce(Sum("vaccine_public"), 0),
                  v_pri=Coalesce(Sum("vaccine_prive"), 0))
    )

    obj_map = {r["maladie_masse"] or "—": int(r["objectif"] or 0) for r in obj_qs}
    vac_map = {r["maladie_masse"] or "—": int((r["v_pub"] or 0) + (r["v_pri"] or 0)) for r in vac_qs}

    maladies = set(obj_map) | set(vac_map)
    ordre_prioritaire = ["PPR", "PE", "MNC", "DNCB", "PPCB"]
    autres = sorted([m for m in maladies if m not in ordre_prioritaire])
    ordered = [m for m in ordre_prioritaire if m in maladies] + autres

    rows = [["Maladie", "Objectifs", "Effectifs vaccinés", "Écart (Obj - Vacc)", "Taux de réalisation (%)"]]
    tot_O = tot_V = 0
    for mal in ordered:
        O = obj_map.get(mal, 0)
        V = vac_map.get(mal, 0)
        E = O - V
        taux = ("∞" if V > 0 else 0) if O == 0 else round((V / O) * 100, 2)
        rows.append([mal, O, V, E, taux])
        tot_O += O; tot_V += V
    tot_E = tot_O - tot_V
    tot_taux = ("∞" if tot_V > 0 else 0) if tot_O == 0 else round((tot_V / tot_O) * 100, 2)
    rows.append(["Total", tot_O, tot_V, tot_E, tot_taux])
    return rows


def t3_evolution(start, end, camp_name: str):
    prev_end = start - date.resolution
    prev_start = prev_end - (end - start)

    def totals(a, b):
        qs = (
            VChild.objects.select_related("parent")
            .filter(parent__campagne__iexact=camp_name,
                    parent__submission_time__date__range=[a, b])
            .values("maladie_masse")
            .annotate(v_pub=Coalesce(Sum("vaccine_public"), 0),
                      v_pri=Coalesce(Sum("vaccine_prive"), 0))
        )
        return {r["maladie_masse"] or "—": int((r["v_pub"] or 0) + (r["v_pri"] or 0)) for r in qs}

    prev = totals(prev_start, prev_end)
    curr = totals(start, end)
    since_start = (
        VChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name,
                parent__submission_time__date__lte=end)
        .values("maladie_masse")
        .annotate(v_pub=Coalesce(Sum("vaccine_public"), 0),
                  v_pri=Coalesce(Sum("vaccine_prive"), 0))
    )
    cumu = {r["maladie_masse"] or "—": int((r["v_pub"] or 0) + (r["v_pri"] or 0)) for r in since_start}

    rows = [["Maladie", "Situation semaine précédente", "Effectifs vaccinés cette période", "Cumul depuis démarrage"]]
    for m in sorted(set(prev) | set(curr) | set(cumu)):
        rows.append([m, prev.get(m, 0), curr.get(m, 0), cumu.get(m, 0)])
    return rows


def _count_calendar_weeks(start_date: date, end_date: date) -> int:
    """Nombre de semaines ISO couvrant l'intervalle [start_date, end_date]."""
    start_monday = start_date - timedelta(days=start_date.weekday())
    end_monday   = end_date   - timedelta(days=end_date.weekday())
    weeks = 0
    cur = start_monday
    while cur <= end_monday:
        weeks += 1
        cur += timedelta(days=7)
    return weeks


def t4_completeness(start, end, camp_name: str):
    rows = [["Régions", "Départements",
             "Rapports attendus par départements depuis le début de la campagne",
             "Rapports reçus", "Complétude (%)"]]

    # Début de campagne (1ère soumission)
    debut = (
        VChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name)
        .annotate(d=TruncDate("parent__submission_time"))
        .order_by("d").values_list("d", flat=True).first()
    )
    debut_campagne = debut or start
    attendus = _count_calendar_weeks(debut_campagne, end)

    # Tous les départements présents pour cette campagne (depuis objectifs OU vaccinations)
    deps_v = (
        VChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name)
        .exclude(parent__grp4_departement__isnull=True).exclude(parent__grp4_departement="")
        .values_list("parent__grp4_departement", "parent__grp4_region").distinct()
    )
    deps_o = (
        OChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name)
        .exclude(parent__grp4_departement__isnull=True).exclude(parent__grp4_departement="")
        .values_list("parent__grp4_departement", "parent__grp4_region").distinct()
    )
    deps = sorted(set(list(deps_v) + list(deps_o)), key=lambda x: (x[1] or "", x[0] or ""))

    for dep, reg in deps:
        recu = (
            VChild.objects.select_related("parent")
            .filter(parent__campagne__iexact=camp_name,
                    parent__grp4_departement__iexact=dep,
                    parent__submission_time__date__gte=debut_campagne,
                    parent__submission_time__date__lte=end)
            .annotate(sem=TruncWeek("parent__submission_time"))
            .values("sem").distinct().count()
        )
        taux = round((recu / attendus) * 100) if attendus else 0
        rows.append([reg or "—", dep or "—", attendus, recu, taux])

    return rows


def t5_real_couv(start, end, camp_name: str):
    obj_qs = (
        OChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name)
        .values("maladie_masse")
        .annotate(
            eligibles=Coalesce(Sum("effectif_elligible"), 0),
            objectifs=Coalesce(Sum("effectif_cible"), 0),
        )
    )
    vac_qs = (
        VChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name,
                parent__submission_time__date__range=[start, end])
        .values("maladie_masse")
        .annotate(v_pub=Coalesce(Sum("vaccine_public"), 0),
                  v_pri=Coalesce(Sum("vaccine_prive"), 0))
    )

    obj_map = {r["maladie_masse"] or "—": r for r in obj_qs}
    vac_map = {r["maladie_masse"] or "—": int((r["v_pub"] or 0) + (r["v_pri"] or 0)) for r in vac_qs}

    priorite = ["PPR", "PE", "MNC", "DNCB", "PPCB"]
    toutes = set(obj_map.keys()) | set(vac_map.keys())
    autres = sorted([m for m in toutes if m not in priorite])
    ordre = [m for m in priorite if m in toutes] + autres

    rows = [["Maladie", "Effectifs éligibles", "Objectifs", "Effectifs vaccinés", "Taux réalisation (%)", "Taux couverture (%)"]]
    tot_elig = tot_obj = tot_vac = 0

    for mal in ordre:
        o = obj_map.get(mal, {"eligibles": 0, "objectifs": 0})
        eli = int((o.get("eligibles") if isinstance(o, dict) else o["eligibles"]) or 0)
        obj = int((o.get("objectifs") if isinstance(o, dict) else o["objectifs"]) or 0)
        vac = int(vac_map.get(mal, 0) or 0)

        taux_real = ("∞" if vac > 0 else 0) if obj == 0 else round((vac / obj) * 100, 2)
        taux_couv = ("∞" if vac > 0 else 0) if eli == 0 else round((vac / eli) * 100, 2)

        rows.append([mal, eli, obj, vac, taux_real, taux_couv])

        tot_elig += eli; tot_obj += obj; tot_vac += vac

    tot_taux_real = ("∞" if tot_vac > 0 else 0) if tot_obj == 0 else round((tot_vac / tot_obj) * 100, 2)
    tot_taux_couv = ("∞" if tot_vac > 0 else 0) if tot_elig == 0 else round((tot_vac / tot_elig) * 100, 2)
    rows.append(["Total", tot_elig, tot_obj, tot_vac, tot_taux_real, tot_taux_couv])
    return rows


def t6_marquage(start, end, camp_name: str):
    qs = (
        VChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name,
                parent__submission_time__date__range=[start, end],
                maladie_masse__iexact="PPR")
        .values("parent__grp4_region")
        .annotate(
            vacc=Coalesce(Sum(F("vaccine_public") + F("vaccine_prive")), 0),
            marque=Coalesce(Sum("total_marque"), 0)
        )
    )

    rows = [["Région", "Effectifs vaccinés", "Effectifs marqués", "Taux de marquage (%)"]]
    total_vacc = total_marque = 0
    for r in qs:
        v = int(r["vacc"] or 0)
        m = int(r["marque"] or 0)
        taux = round((m / v) * 100, 2) if v else 0
        rows.append([r["parent__grp4_region"] or "—", v, m, taux])
        total_vacc += v; total_marque += m

    taux_nat = round((total_marque / total_vacc) * 100, 2) if total_vacc else 0
    rows.append(["National", total_vacc, total_marque, taux_nat])
    return rows

def t7_public_prive(start, end, camp_name: str):
    """
    T7 (par département) :
    [Région, Département, Maladie, Secteur public, Secteur privé, Total]

    - Filtre sur la campagne (nom exact, insensible à la casse) et la période (date de submission parent).
    - Agrégation sur les champs Kobo : vaccine_public / vaccine_prive.
    """
    qs = (
        VChild.objects.select_related("parent")
        .filter(
            parent__campagne__iexact=camp_name,
            parent__submission_time__date__range=[start, end],
        )
        .values("parent__grp4_region", "parent__grp4_departement", "maladie_masse")
        .annotate(
            pub=Coalesce(Sum("vaccine_public"), 0),
            pri=Coalesce(Sum("vaccine_prive"), 0),
        )
    )

    # Tri Python (gère proprement les None)
    data = sorted(
        qs,
        key=lambda r: (
            (r["parent__grp4_region"] or ""),
            (r["parent__grp4_departement"] or ""),
            (r["maladie_masse"] or ""),
        ),
    )

    rows = [["Région", "Département", "Maladie", "Secteur public", "Secteur privé", "Total"]]
    total_pub = total_pri = 0

    for r in data:
        reg = r["parent__grp4_region"] or "—"
        dep = r["parent__grp4_departement"] or "—"
        mal = r["maladie_masse"] or "—"
        pub = int(r["pub"] or 0)
        pri = int(r["pri"] or 0)
        rows.append([reg, dep, mal, pub, pri, pub + pri])
        total_pub += pub
        total_pri += pri

    # ligne de synthèse nationale
    rows.append(["National", "—", "—", total_pub, total_pri, total_pub + total_pri])
    return rows



def t8_yoy(start, end, camp_name: str):
    """Comparaison période N vs même période N-1, pour la même campagne (si données existent)."""
    prev_start = _shift_year_safe(start, -1)
    prev_end   = _shift_year_safe(end, -1)

    this = (
        VChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name,
                parent__submission_time__date__range=[start, end])
        .values("maladie_masse")
        .annotate(v_pub=Coalesce(Sum("vaccine_public"), 0),
                  v_pri=Coalesce(Sum("vaccine_prive"), 0))
    )
    prev = (
        VChild.objects.select_related("parent")
        .filter(parent__campagne__iexact=camp_name,
                parent__submission_time__date__range=[prev_start, prev_end])
        .values("maladie_masse")
        .annotate(v_pub=Coalesce(Sum("vaccine_public"), 0),
                  v_pri=Coalesce(Sum("vaccine_prive"), 0))
    )

    this_map = {r["maladie_masse"] or "—": int((r["v_pub"] or 0) + (r["v_pri"] or 0)) for r in this}
    prev_map = {r["maladie_masse"] or "—": int((r["v_pub"] or 0) + (r["v_pri"] or 0)) for r in prev}

    rows = [["Maladie", f"Situation du {_shift_year_safe(start, -1).strftime('%d/%m/%Y')}",
             "Écarts", f"Situation du {end.strftime('%d/%m/%Y')}"]]
    for m in sorted(set(this_map) | set(prev_map)):
        a = prev_map.get(m, 0); b = this_map.get(m, 0)
        rows.append([m, a, b - a, b])
    return rows


def t9_ciblees(start, end, camp_name: str):
    """
    Vaccinations ciblées : si la campagne choisie est de type 'Ciblée' (ou 'Ciblee'), 
    on filtre dessus, sinon on prend toutes les campagnes de type 'Ciblée' sur la période.
    """
    # Détecter le type de la campagne choisie (dans ObjectifSn ou VaccinationSn)
    types = list(
        ObjectifSn.objects.filter(campagne__iexact=camp_name)
                          .values_list("type_de_campagne", flat=True).distinct()
    ) + list(
        VaccinationSn.objects.filter(campagne__iexact=camp_name)
                             .values_list("type_de_campagne", flat=True).distinct()
    )
    types_norm = { (t or "").strip().lower() for t in types }

    qs = VChild.objects.select_related("parent").filter(
        parent__submission_time__date__range=[start, end]
    )
    if {"ciblée", "ciblee"} & types_norm:
        qs = qs.filter(parent__campagne__iexact=camp_name)
    else:
        qs = qs.filter(parent__type_de_campagne__iregex=r"^cibl[ée]e$")

    agg = (
        qs.values("maladie_masse")
          .annotate(v=Coalesce(Sum(F("vaccine_public") + F("vaccine_prive")), 0))
          .order_by("maladie_masse")
    )

    rows = [["Maladie ciblée", "Effectifs vaccinés"]]
    total = 0
    for r in agg:
        val = int(r["v"] or 0)
        rows.append([r["maladie_masse"] or "—", val])
        total += val
    rows.append(["National", total])
    return rows


# ----------------------------- Vue principale -----------------------------
def camvac_excel(request):
    form = RapportCamvacFilterForm(request.GET or None)
    if not form.is_valid():
        return render(request, "vaccination_sn/report_form.html", {"form": form})

    camp_name = form.cleaned_data["campagne"]
    start, end, periode_label = resolve_period(form.cleaned_data)

    sheets = {
        "T1 - Situation régionale":      t1_situation_regionale(start, end, camp_name),
        "T2 - Écarts nationaux":         t2_ecarts_nationaux(start, end, camp_name),
        "T3 - Évolution":                t3_evolution(start, end, camp_name),
        "T4 - Complétude":               t4_completeness(start, end, camp_name),
        "T5 - Réalisation & Couverture": t5_real_couv(start, end, camp_name),
        "T6 - Marquage PPR":             t6_marquage(start, end, camp_name),
        "T8 - Comparaison N vs N-1":     t8_yoy(start, end, camp_name),
        "T9 - Vaccinations ciblées":     t9_ciblees(start, end, camp_name),
    }

    # T7 : on a déjà public/privé explicite côté Kobo — on l’ajoute systématiquement
    sheets["T7 - Public vs Privé (par département)"] = t7_public_prive(start, end, camp_name)


    xlsx_bytes = write_sheets(sheets)
    safe_camp = "".join(c for c in camp_name if c.isalnum() or c in ("-", "_")).strip() or "campagne"
    filename = f"Rapport_CAMVAC_{safe_camp}_{periode_label}.xlsx"

    resp = HttpResponse(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

